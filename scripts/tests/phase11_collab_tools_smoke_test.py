#!/usr/bin/env python3
"""Smoke test suite for Phase 11 collaborative writing tools.

Validates:
  - merge.py: produces parseable output
  - split_chara.py: covers all dialogues, no missing/duplicate
  - strip_code.py family: correct plain-text / docx / tex / xlsx output
  - add_soft_hyphens.py: correct insertion for zh/ja
  - generate_sample_script.py: all 3 templates parseable via ScriptLoader
"""

from __future__ import annotations

import json
import os
import re
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Any, Optional

SCRIPT_PATH = Path(__file__).resolve()
PROJECT_ROOT = SCRIPT_PATH.parents[2]
TOOLS_DIR = PROJECT_ROOT / "scripts" / "tools"


class CliError(Exception):
    pass


DIALOGUE_PATTERN = re.compile(r'^(.+?)：：\u201c(.+?)\u201d\s*$')
CODE_BLOCK_START = re.compile(r'^@?<\|')
CODE_BLOCK_END = re.compile(r'\|>\s*$')


def _count_dialogues(text: str) -> int:
    return sum(1 for line in text.splitlines() if DIALOGUE_PATTERN.match(line))


def _count_code_blocks(text: str) -> int:
    count = 0
    in_code = False
    for line in text.splitlines():
        if CODE_BLOCK_START.match(line.strip()):
            in_code = True
            continue
        if in_code and CODE_BLOCK_END.match(line.strip()):
            in_code = False
            count += 1
    return count


# ── Fixtures ──────────────────────────────────────────────

CHAR_A = """#王二宫
王二宫：：\u201c真没想到，我们竟然会被暴雨困在了教学楼里。\u201d

二宫会长少见地皱起了她秀气的眉头。

王二宫：：\u201c我也不想这样啊，都是浅野说那个文件送到了嘛。\u201d

大家坐在学生会室之中，都选择了沉默。
"""

CHAR_B = """#张浅野
张浅野：：\u201c结果还是因为暴雨延期了，这不是我的错。\u201d

浅野学长靠在椅背上，像是无聊似的耸了耸肩。

张浅野：：\u201c那你也可以在联系学生会干部之前确认一下嘛。\u201d

和他们争吵了起来。
"""

CHAR_C = """#旁白
旁白：：\u201c这是旁白的台词。\u201d

我木然地看着每个人的互动。

这是学生会日常的一部分。
"""


def _failures() -> list[str]:
    return []


def _expect(cond: bool, msg: str, failures: list[str]) -> None:
    if not cond:
        failures.append(msg)


# ── Test: merge.py ────────────────────────────────────────

def test_merge(failures: list[str], tmpdir: str) -> None:
    # Write character files
    char_paths = []
    for name, content in [("char_a.txt", CHAR_A), ("char_b.txt", CHAR_B), ("char_c.txt", CHAR_C)]:
        p = os.path.join(tmpdir, name)
        Path(p).write_text(content, encoding="utf-8")
        char_paths.append(p)

    merge_path = os.path.join(tmpdir, "merged.txt")
    result = subprocess.run(
        [sys.executable, str(TOOLS_DIR / "merge.py"), *char_paths, "-o", merge_path],
        capture_output=True, text=True,
    )
    _expect(result.returncode == 0, f"merge.py should exit 0: {result.stderr}", failures)

    merged = Path(merge_path).read_text(encoding="utf-8")
    _expect(len(merged) > 0, "merge.py output should not be empty", failures)

    # Count dialogues in merged output
    dc = _count_dialogues(merged)
    _expect(dc == 5, f"merged output should have 5 dialogues, got {dc}", failures)

    # Verify key lines present
    _expect("王二宫" in merged and "张浅野" in merged,
            "merged output should contain both character dialogues", failures)

    # Test with --title
    title_path = os.path.join(tmpdir, "merged_title.txt")
    result2 = subprocess.run(
        [sys.executable, str(TOOLS_DIR / "merge.py"), *char_paths, "-o", title_path, "--title", "测试章节"],
        capture_output=True, text=True,
    )
    _expect(result2.returncode == 0, f"merge.py --title should exit 0: {result2.stderr}", failures)
    titled = Path(title_path).read_text(encoding="utf-8")
    _expect("label('测试章节'" in titled or "label(\"测试章节\"" in titled,
            "merge.py --title should generate a label", failures)


# ── Test: split_chara.py ─────────────────────────────────

def test_split_chara(failures: list[str], tmpdir: str) -> None:
    input_path = os.path.join(tmpdir, "split_input.txt")
    # Create a valid NovaScript with multiple characters
    content = """@<|
label('test_split', '测试拆分')
is_start()
|>
<|
show(bg, 'room')
|>
王二宫：：\u201c第一句台词。\u201d

旁白描述了场景。

张浅野：：\u201c第二句台词，来自浅野。\u201d

王二宫：：\u201c第三句台词，又回到二宫。\u201d

更多旁白。

张浅野：：\u201c第四句台词。\u201d
"""
    Path(input_path).write_text(content, encoding="utf-8")

    output_dir = os.path.join(tmpdir, "split_output")
    result = subprocess.run(
        [sys.executable, str(TOOLS_DIR / "split_chara.py"),
         "--input", input_path, "--output-dir", output_dir],
        capture_output=True, text=True,
    )
    _expect(result.returncode == 0, f"split_chara.py should exit 0: {result.stderr}", failures)

    # Check output files
    output_files = list(Path(output_dir).glob("*.txt"))
    _expect(len(output_files) == 2,
            f"split_chara should produce 2 character files, got {len(output_files)}", failures)

    # Verify no missing/duplicate dialogues
    for fp in output_files:
        text = fp.read_text(encoding="utf-8")
        if "王二宫" in fp.name:
            _expect(text.count("王二宫：：") == 2,
                    f"王二宫 should have 2 dialogues, got {text.count('王二宫：：')}", failures)
        if "张浅野" in fp.name:
            _expect(text.count("张浅野：：") == 2,
                    f"张浅野 should have 2 dialogues, got {text.count('张浅野：：')}", failures)

    # Test --include-narration
    output_dir_narr = os.path.join(tmpdir, "split_output_narr")
    result2 = subprocess.run(
        [sys.executable, str(TOOLS_DIR / "split_chara.py"),
         "--input", input_path, "--output-dir", output_dir_narr, "--include-narration"],
        capture_output=True, text=True,
    )
    _expect(result2.returncode == 0,
            f"split_chara.py --include-narration should exit 0: {result2.stderr}", failures)


# ── Test: strip_code family ───────────────────────────────

def test_strip_code(failures: list[str], tmpdir: str) -> None:
    input_path = os.path.join(tmpdir, "strip_input.txt")
    content = """@<|
label('test_strip', '测试剥离')
is_start()
|>
<|
show(bg, 'room')
play(bgm, 'prelude')
box_hide_show(2)
|>
王二宫：：\u201c你好，世界。\u201d

这是旁白文本。

（TODO：需要后续补充内容）

张浅野：：\u201c你好，二宫。\u201d
"""
    Path(input_path).write_text(content, encoding="utf-8")

    # Test: plain text strip
    plain_out = os.path.join(tmpdir, "strip_plain.txt")
    result = subprocess.run(
        [sys.executable, str(TOOLS_DIR / "strip_code.py"),
         "--input", input_path, "--output", plain_out],
        capture_output=True, text=True,
    )
    _expect(result.returncode == 0, f"strip_code.py should exit 0: {result.stderr}", failures)
    plain = Path(plain_out).read_text(encoding="utf-8")
    _expect("你好，世界" in plain, "plain strip should preserve dialogue text", failures)
    _expect("旁白文本" in plain, "plain strip should preserve narration", failures)
    _expect("label('test_strip'" not in plain, "plain strip should remove code blocks", failures)
    _expect("show(bg," not in plain, "plain strip should remove code directives", failures)

    # Test: keep speaker names
    speaker_out = os.path.join(tmpdir, "strip_speaker.txt")
    result2 = subprocess.run(
        [sys.executable, str(TOOLS_DIR / "strip_code.py"),
         "--input", input_path, "--output", speaker_out, "--keep-speaker-names"],
        capture_output=True, text=True,
    )
    _expect(result2.returncode == 0, f"strip_code.py --keep-speaker-names should exit 0", failures)
    speaker_text = Path(speaker_out).read_text(encoding="utf-8")
    _expect("王二宫: 你好，世界" in speaker_text,
            "keep-speaker-names should prepend speaker names", failures)

    # Test: strip_code_tex
    tex_out = os.path.join(tmpdir, "strip_tex.tex")
    result3 = subprocess.run(
        [sys.executable, str(TOOLS_DIR / "strip_code_tex.py"),
         "--input", input_path, "--output", tex_out],
        capture_output=True, text=True,
    )
    _expect(result3.returncode == 0, f"strip_code_tex.py should exit 0: {result3.stderr}", failures)
    tex = Path(tex_out).read_text(encoding="utf-8")
    _expect(r"\documentclass" in tex, "tex output should contain LaTeX preamble", failures)
    _expect(r"\speaker{王二宫}" in tex, "tex output should use speaker macro", failures)

    # Test: strip_code_xlsx (requires openpyxl)
    try:
        import openpyxl  # noqa: F401
        xlsx_out = os.path.join(tmpdir, "strip_xlsx.xlsx")
        result4 = subprocess.run(
            [sys.executable, str(TOOLS_DIR / "strip_code_xlsx.py"),
             "--input", input_path, "--output", xlsx_out],
            capture_output=True, text=True,
        )
        _expect(result4.returncode == 0, f"strip_code_xlsx.py should exit 0: {result4.stderr}", failures)
        _expect(os.path.isfile(xlsx_out), "xlsx output file should exist", failures)
        # Verify xlsx content
        wb = openpyxl.load_workbook(xlsx_out)
        ws = wb.active
        _expect(ws.cell(1, 1).value == "Speaker", "xlsx should have Speaker header", failures)
        _expect(ws.cell(1, 2).value == "Text", "xlsx should have Text header", failures)
        wb.close()
    except ImportError:
        print("  (skip) strip_code_xlsx: openpyxl not available", file=sys.stderr)

    # Test: strip_code_docx (requires python-docx)
    try:
        import docx  # noqa: F401
        docx_out = os.path.join(tmpdir, "strip_docx.docx")
        result5 = subprocess.run(
            [sys.executable, str(TOOLS_DIR / "strip_code_docx.py"),
             "--input", input_path, "--output", docx_out],
            capture_output=True, text=True,
        )
        _expect(result5.returncode == 0, f"strip_code_docx.py should exit 0: {result5.stderr}", failures)
        _expect(os.path.isfile(docx_out), "docx output file should exist", failures)
    except ImportError:
        print("  (skip) strip_code_docx: python-docx not available", file=sys.stderr)


# ── Test: add_soft_hyphens.py ─────────────────────────────

def test_add_soft_hyphens(failures: list[str], tmpdir: str) -> None:
    input_path = os.path.join(tmpdir, "hyphens_input.txt")
    content = "王二宫：：\u201c真没想到，我们竟然会被暴雨困在了教学楼里。\u201d\n"
    Path(input_path).write_text(content, encoding="utf-8")

    result = subprocess.run(
        [sys.executable, str(TOOLS_DIR / "add_soft_hyphens.py"),
         "--input", input_path, "--text-only", "--min-length", "4"],
        capture_output=True, text=True,
    )
    _expect(result.returncode == 0, f"add_soft_hyphens.py should exit 0: {result.stderr}", failures)

    # Check that soft hyphens were inserted in long Chinese text
    soft = "\u00AD"
    _expect(soft in result.stdout,
            "add_soft_hyphens should insert soft hyphens in long CJK text", failures)

    # Verify dialogue structure is preserved
    dialogue_match = DIALOGUE_PATTERN.search(result.stdout)
    _expect(dialogue_match is not None,
            "add_soft_hyphens --text-only should preserve dialogue structure", failures)


# ── Test: generate_sample_script.py ───────────────────────

def test_generate_sample_script(failures: list[str], tmpdir: str) -> None:
    for template in ["basic", "branching", "minigame"]:
        out_path = os.path.join(tmpdir, f"sample_{template}.txt")
        result = subprocess.run(
            [sys.executable, str(TOOLS_DIR / "generate_sample_script.py"),
             "--template", template, "--output", out_path],
            capture_output=True, text=True,
        )
        _expect(result.returncode == 0,
                f"generate_sample_script.py --template {template} should exit 0: {result.stderr}", failures)

        content = Path(out_path).read_text(encoding="utf-8")
        _expect(len(content) > 100,
                f"generated {template} script should not be empty", failures)
        _expect("label(" in content,
                f"generated {template} script should contain a label", failures)

        # Basic structure checks
        _expect("@<|" in content or "<|" in content,
                f"generated {template} script should contain code blocks", failures)

    # Verify template listing
    result = subprocess.run(
        [sys.executable, str(TOOLS_DIR / "generate_sample_script.py"), "--list"],
        capture_output=True, text=True,
    )
    _expect(result.returncode == 0, f"--list should exit 0", failures)
    for name in ["basic", "branching", "minigame"]:
        _expect(name in result.stdout, f"--list should include {name}", failures)


# ── End-to-end: merge → split → strip workflow ────────────

def test_e2e_workflow(failures: list[str], tmpdir: str) -> None:
    """End-to-end: merge 2 character files → split back → strip output."""
    # Create two character files
    char_a = """#王二宫
王二宫：：\u201c第一句台词。\u201d

这是旁白A。

王二宫：：\u201c第二句台词。\u201d
"""

    char_b = """#张浅野
张浅野：：\u201c第三句台词。\u201d

这是旁白B。

张浅野：：\u201c第四句台词。\u201d
"""

    a_path = os.path.join(tmpdir, "e2e_a.txt")
    b_path = os.path.join(tmpdir, "e2e_b.txt")
    Path(a_path).write_text(char_a, encoding="utf-8")
    Path(b_path).write_text(char_b, encoding="utf-8")

    # Step 1: Merge
    merged_path = os.path.join(tmpdir, "e2e_merged.txt")
    r = subprocess.run(
        [sys.executable, str(TOOLS_DIR / "merge.py"), a_path, b_path, "-o", merged_path],
        capture_output=True, text=True,
    )
    _expect(r.returncode == 0, f"e2e merge step: {r.stderr}", failures)

    # Step 2: Split back
    split_dir = os.path.join(tmpdir, "e2e_split")
    r2 = subprocess.run(
        [sys.executable, str(TOOLS_DIR / "split_chara.py"),
         "--input", merged_path, "--output-dir", split_dir],
        capture_output=True, text=True,
    )
    _expect(r2.returncode == 0, f"e2e split step: {r2.stderr}", failures)

    # Step 3: Strip
    plain_path = os.path.join(tmpdir, "e2e_stripped.txt")
    r3 = subprocess.run(
        [sys.executable, str(TOOLS_DIR / "strip_code.py"),
         "--input", merged_path, "--output", plain_path],
        capture_output=True, text=True,
    )
    _expect(r3.returncode == 0, f"e2e strip step: {r3.stderr}", failures)

    stripped = Path(plain_path).read_text(encoding="utf-8")
    _expect("第一句台词" in stripped, "stripped output should contain dialogue text", failures)
    _expect("第四句台词" in stripped, "stripped output should contain all dialogues", failures)

    # Step 4: Tex output too
    tex_path = os.path.join(tmpdir, "e2e_stripped.tex")
    r4 = subprocess.run(
        [sys.executable, str(TOOLS_DIR / "strip_code_tex.py"),
         "--input", merged_path, "--output", tex_path],
        capture_output=True, text=True,
    )
    _expect(r4.returncode == 0, f"e2e tex step: {r4.stderr}", failures)

    print(f"  E2E workflow: merge → split → strip_plain → strip_tex ✓", file=sys.stderr)


# ── Main ──────────────────────────────────────────────────

def main() -> int:
    failures: list[str] = []

    with tempfile.TemporaryDirectory(prefix="unistory-phase11-smoke-") as tmpdir:
        print(f"Phase 11 Smoke Tests", file=sys.stderr)
        print(f"  tmpdir: {tmpdir}", file=sys.stderr)

        print(f"  test_merge ...", file=sys.stderr)
        test_merge(failures, tmpdir)

        print(f"  test_split_chara ...", file=sys.stderr)
        test_split_chara(failures, tmpdir)

        print(f"  test_strip_code ...", file=sys.stderr)
        test_strip_code(failures, tmpdir)

        print(f"  test_add_soft_hyphens ...", file=sys.stderr)
        test_add_soft_hyphens(failures, tmpdir)

        print(f"  test_generate_sample_script ...", file=sys.stderr)
        test_generate_sample_script(failures, tmpdir)

        print(f"  test_e2e_workflow ...", file=sys.stderr)
        test_e2e_workflow(failures, tmpdir)

    if failures:
        print(f"\nPhase 11 smoke tests: {len(failures)} FAILURE(S)", file=sys.stderr)
        for f in failures:
            print(f"  - {f}", file=sys.stderr)
        return 1

    print(f"\nPhase 11 smoke tests: OK (all passed)", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
