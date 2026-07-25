#!/usr/bin/env python3
"""strip_code_xlsx.py — strip NovaScript code blocks and output as Excel (.xlsx).

Output columns: Speaker | Text | Line | File
Useful for translation workflows and content review spreadsheets.

Requires: openpyxl (`pip install openpyxl`)
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Optional

SCRIPT_PATH = Path(__file__).resolve()
PROJECT_ROOT = SCRIPT_PATH.parents[2]


class CliError(Exception):
    """An invocation or infrastructure error that returns exit code 2."""


DIALOGUE_PATTERN = re.compile(r'^(.+?)：：\u201c(.+?)\u201d\s*$')
CODE_BLOCK_START = re.compile(r'^@?<\|')
CODE_BLOCK_END = re.compile(r'\|>\s*$')


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Strip NovaScript code blocks and output as Excel (.xlsx).",
    )
    parser.add_argument(
        "--input", "-i", metavar="PATH", required=True,
        help="input NovaScript scenario file",
    )
    parser.add_argument(
        "--output", "-o", metavar="PATH", required=True,
        help="output .xlsx file path",
    )
    parser.add_argument(
        "--project", metavar="PATH",
        help="project root (defaults to repository root)",
    )
    return parser


def resolve_project_root(explicit: str | None) -> Path:
    if explicit:
        p = Path(explicit).resolve()
        if not p.is_dir():
            raise CliError(f"project root not found: {p}")
        return p
    return PROJECT_ROOT


def strip_to_xlsx(filepath: Path, output: Path) -> dict:
    """Strip code and write to Excel, return counts."""
    try:
        raw = filepath.read_text(encoding="utf-8")
    except Exception as exc:
        raise CliError(f"cannot read {filepath}: {exc}") from exc

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise CliError(
            "openpyxl is required. Install it with: pip install openpyxl"
        )

    lines = raw.splitlines()
    wb = Workbook()
    ws = wb.active
    ws.title = filepath.stem[:31]  # Sheet name max 31 chars

    # Headers
    headers = ["Speaker", "Text", "Line", "File"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row = 2
    in_code = False
    dialogue_count = 0
    narration_count = 0

    for lineno, line in enumerate(lines, start=1):
        stripped = line.strip()

        if CODE_BLOCK_START.match(stripped):
            in_code = True
            continue
        if in_code:
            if CODE_BLOCK_END.match(stripped):
                in_code = False
            continue

        if not stripped:
            continue

        m = DIALOGUE_PATTERN.match(stripped)
        if m:
            speaker = m.group(1).strip()
            text = m.group(2).strip()
            ws.cell(row=row, column=1, value=speaker)
            ws.cell(row=row, column=2, value=text)
            ws.cell(row=row, column=3, value=lineno)
            ws.cell(row=row, column=4, value=filepath.name)
            dialogue_count += 1
        else:
            ws.cell(row=row, column=1, value="[Narration]")
            ws.cell(row=row, column=2, value=stripped)
            ws.cell(row=row, column=3, value=lineno)
            ws.cell(row=row, column=4, value=filepath.name)
            # Style narration differently
            narration_font = Font(italic=True, color="666666")
            for col in range(1, 5):
                ws.cell(row=row, column=col).font = narration_font
            narration_count += 1
        row += 1

    # Adjust column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 25

    # Freeze header row
    ws.freeze_panes = "A2"

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output))

    return {"dialogue_count": dialogue_count, "narration_count": narration_count, "total_rows": row - 2}


def main() -> None:
    args = _parser().parse_args()
    try:
        project_root = resolve_project_root(args.project)
    except CliError as e:
        print(str(e), file=sys.stderr)
        sys.exit(2)

    input_path = Path(args.input)
    if not input_path.is_absolute():
        cwd_fp = Path.cwd() / input_path
        proj_fp = project_root / input_path
        input_path = cwd_fp if cwd_fp.exists() else proj_fp
    if not input_path.is_file():
        print(f"strip_code_xlsx: error: input file not found: {input_path}", file=sys.stderr)
        sys.exit(2)

    output_path = Path(args.output)
    if not output_path.is_absolute():
        output_path = Path.cwd() / output_path

    try:
        report = strip_to_xlsx(input_path, output_path)
        print(f"Written: {output_path}")
        print(f"  Dialogues: {report['dialogue_count']}")
        print(f"  Narrations: {report['narration_count']}")
        print(f"  Total rows: {report['total_rows']}")
    except CliError as e:
        print(f"strip_code_xlsx: error: {e}", file=sys.stderr)
        sys.exit(2)


if __name__ == "__main__":
    main()
